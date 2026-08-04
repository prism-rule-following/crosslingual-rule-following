#!/usr/bin/env python3
"""
validate_dataset.py
===================

Reproducible integrity validator for the cross-lingual rule-following dataset.
Runs against a LOCAL clone where all 13 language files are
intact -- NOT against GitHub URLs (several of which are broken, and the files
are too large to fetch whole anyway).


What it checks (the seven issues, each emitted as its own issue-group)
---------------------------------------------------------------------
  1. identical_contrastive_pair : system_rule == system_non_rule (zero signal)
  2. missing_pair_type          : no pair_type tag (split: otherwise-well-formed vs. also-issue-1)
  3. translation_drift          : per-row diff vs English base -- flags rows where the
                                  contrastive suffix that EXISTS in English has collapsed
                                  or gone missing after translation (the real root cause)
  4. llm_judge_checker          : checker is free-text / "llm-judge" rather than executable;
                                  reports which categories are 100% llm-judge, separating the
                                  documented-intentional set from active_cancelled (undocumented)
  5. custom_checker_dependency  : checker calls check_second_word / check_single_word
                                  (fine, but only runs if adherence_scoring.py is importable)
  6. duplicate_rows             : identical (system_rule, user_query, category) groups
  7. duplicate_ids              : repeated id values (should be zero)

Plus schema checks: missing required fields, and row-count vs the English base.

Output
------
An .xlsx workbook, one sheet per language (+ a SUMMARY sheet), with rows GROUPED
by issue type and the affected row-IDs listed per group -- so the team can triage
by issue rather than scrolling raw rows. A machine-readable JSON report is written
alongside it for CI use.

Usage
-----
    python validate_dataset.py --data-dir canonical/data --out validation_report.xlsx

    # CI gate: exit non-zero if any blocking issue is present
    python validate_dataset.py --data-dir canonical/data --ci

Dependencies: openpyxl (preinstalled). Only stdlib otherwise.
"""

import argparse
import json
import os
import sys
from collections import defaultdict, Counter

# ---------------------------------------------------------------------------
# Configuration -- edit here if the dataset's conventions change.
# ---------------------------------------------------------------------------

ENGLISH_FILENAME = "full_dataset.json"

# ISO code -> filename suffix. English base has no suffix.
LANGUAGE_FILES = {
    "en": "full_dataset.json",
    "am": "full_dataset_am.json",
    "de": "full_dataset_de.json",
    "hi": "full_dataset_hi.json",
    "ig": "full_dataset_ig.json",
    "it": "full_dataset_it.json",
    "ko": "full_dataset_ko.json",
    "ru": "full_dataset_ru.json",
    "sw": "full_dataset_sw.json",
    "ta": "full_dataset_ta.json",
    "tr": "full_dataset_tr.json",
    "ur": "full_dataset_ur.json",
    "yo": "full_dataset_yo.json",
}

REQUIRED_FIELDS = [
    "id",
    "category",
    "language",
    "system_rule",
    "system_non_rule",
    "user_query",
    "checker",
    "rule_clause",
]

# Categories documented in the dataset's own metadata as intentionally
# 100%-LLM-judge (no regex exists for these judgments). active_cancelled is
# deliberately NOT in this set -- it is the one that looks like it should have
# an executable checker but doesn't, per the team's issue #4.
DOCUMENTED_LLM_JUDGE_CATEGORIES = {
    "tone_provocation",
    "humor",
    "directness",
    "humility",
    "emotional_expressiveness",
}

# Keep in sync with DataCategories in canonical/data/model/dataset.py.
KNOWN_CATEGORIES = {
    "ack_invert",
    "active_cancelled",
    "banned_word",
    "bold_html",
    "directness",
    "emotional_expressiveness",
    "humility",
    "humor",
    "include_word",
    "language",
    "second_word",
    "single_word",
    "start_with",
    "tone_provocation",
    "word_count",
}

# A checker is treated as non-executable ("llm-judge") only if it matches this
# exact prefix -- a strict prefix check, not a loose substring match, so a
# checker that happens to contain the word "manual" elsewhere isn't
# misclassified. Missing checkers are handled by check_missing_checker
# instead of being folded in here.
LLM_JUDGE_PREFIX = "manual/llm-judge"

CUSTOM_CHECKER_FNS = ("check_second_word", "check_single_word")


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_pairs(path):
    """Return the list of row dicts from a dataset file, tolerating both
    {'pairs': [...]} and a bare [...] top level."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "pairs" in data:
        return data["pairs"], data.get("metadata", {})
    if isinstance(data, list):
        return data, {}
    raise ValueError(
        f"{path}: unrecognized top-level structure "
        f"(expected list or object with 'pairs')"
    )


# ---------------------------------------------------------------------------
# Per-file checks. Each returns a dict: issue_name -> {"ids": [...], "note": str, "blocking": bool}
# ---------------------------------------------------------------------------


def check_identical_contrastive_pair(pairs):
    ids = [
        r.get("id")
        for r in pairs
        if r.get("system_rule") is not None
        and r.get("system_rule") == r.get("system_non_rule")
    ]
    return {
        "ids": ids,
        "note": "system_rule == system_non_rule verbatim -- row carries zero contrastive "
        "signal. In English this never happens (differing Rule status suffix); "
        "presence here indicates translation collapsed the distinction.",
        "blocking": True,
    }


def check_missing_pair_type(pairs):
    identical = {
        r.get("id") for r in pairs if r.get("system_rule") == r.get("system_non_rule")
    }
    missing_and_broken, missing_but_wellformed = [], []
    for r in pairs:
        if not r.get("pair_type"):
            if r.get("id") in identical:
                missing_and_broken.append(r.get("id"))
            else:
                missing_but_wellformed.append(r.get("id"))
    return {
        "ids": missing_but_wellformed + missing_and_broken,
        "ids_wellformed_metadata_only": missing_but_wellformed,
        "ids_also_identical_pair": missing_and_broken,
        "note": f"Missing pair_type. {len(missing_but_wellformed)} are otherwise "
        f"well-formed (pure metadata backfill -- easy fix); "
        f"{len(missing_and_broken)} overlap with the identical-pair defect "
        f"(issue 1) and need content regeneration, not just a tag.",
        "blocking": False,
    }


def check_translation_drift(pairs, english_by_id):
    """The root-cause check. For each row that HAS an English counterpart, compare
    whether the contrastive structure survived. Flags a row when English had a
    differing rule/non_rule pair but the translated row has them identical, OR when
    English carried a pair_type the translation dropped."""
    if not english_by_id:
        return {
            "ids": [],
            "note": "No English base supplied -- drift check skipped.",
            "blocking": False,
        }
    drift_collapsed_pair, drift_dropped_pairtype = [], []
    for r in pairs:
        rid = r.get("id")
        en = english_by_id.get(rid)
        if en is None:
            continue
        en_differs = en.get("system_rule") != en.get("system_non_rule")
        tr_differs = r.get("system_rule") != r.get("system_non_rule")
        if en_differs and not tr_differs:
            drift_collapsed_pair.append(rid)
        if en.get("pair_type") and not r.get("pair_type"):
            drift_dropped_pairtype.append(rid)
    all_ids = sorted(set(drift_collapsed_pair) | set(drift_dropped_pairtype))
    return {
        "ids": all_ids,
        "ids_collapsed_pair": drift_collapsed_pair,
        "ids_dropped_pairtype": drift_dropped_pairtype,
        "note": f"Translation-introduced defects vs English base: "
        f"{len(drift_collapsed_pair)} rows where a contrastive pair that DIFFERS "
        f"in English became identical after translation; "
        f"{len(drift_dropped_pairtype)} where pair_type present in English was dropped.",
        "blocking": True,
    }


def check_llm_judge_checkers(pairs):
    by_cat_total = Counter()
    by_cat_llmjudge = Counter()
    for r in pairs:
        cat = r.get("category")
        by_cat_total[cat] += 1
        checker = str(r.get("checker") or "").strip().lower()
        if checker.startswith(LLM_JUDGE_PREFIX):
            by_cat_llmjudge[cat] += 1
    fully_llm = {
        c
        for c in by_cat_total
        if by_cat_llmjudge[c] == by_cat_total[c] and by_cat_total[c] > 0
    }
    documented = sorted(fully_llm & DOCUMENTED_LLM_JUDGE_CATEGORIES)
    undocumented = sorted(fully_llm - DOCUMENTED_LLM_JUDGE_CATEGORIES)
    ids_undocumented = [r.get("id") for r in pairs if r.get("category") in undocumented]
    return {
        "ids": ids_undocumented,
        "categories_fully_llm_judge_documented": documented,
        "categories_fully_llm_judge_UNDOCUMENTED": undocumented,
        "note": f"100%-LLM-judge categories (checker starts with '{LLM_JUDGE_PREFIX}'). "
        f"Documented-intentional: {documented or 'none'}. "
        f"UNDOCUMENTED (looks like it should have an executable checker): "
        f"{undocumented or 'none'} -- resolve by design before submission.",
        "blocking": False,
    }


def check_missing_checker(pairs):
    """Separate from check_llm_judge_checkers: a genuinely absent checker is a
    blocking defect (the row can't be evaluated at all), not a stylistic
    'llm-judge vs executable' classification question."""
    ids = [r.get("id") for r in pairs if not str(r.get("checker") or "").strip()]
    return {
        "ids": ids,
        "note": (
            f"{len(ids)} rows have no checker at all (neither executable nor LLM-judge text)."
            if ids
            else "Every row has a checker (executable or LLM-judge)."
        ),
        "blocking": True,
    }


def check_category_validity(pairs):
    bad = [r for r in pairs if r.get("category") not in KNOWN_CATEGORIES]
    ids = [r.get("id") for r in bad]
    bad_values = Counter(r.get("category") for r in bad)
    return {
        "ids": ids,
        "note": (
            f"{len(ids)} rows have a category not in the known set: {dict(bad_values)}"
            if ids
            else "All category values are recognized."
        ),
        "blocking": True,
    }


def check_custom_checker_dependency(pairs):
    ids = [
        r.get("id")
        for r in pairs
        if any(fn in str(r.get("checker", "")) for fn in CUSTOM_CHECKER_FNS)
    ]
    return {
        "ids": ids,
        "note": "Checker calls check_second_word/check_single_word -- valid, but only "
        "executes if adherence_scoring.py is importable wherever checkers run. "
        "Dependency to document, not a defect.",
        "blocking": False,
    }


def check_duplicate_rows(pairs):
    seen = defaultdict(list)
    for r in pairs:
        key = (r.get("system_rule"), r.get("user_query"), r.get("category"))
        seen[key].append(r.get("id"))
    dup_groups = {k: v for k, v in seen.items() if len(v) > 1}
    flat_ids = [rid for v in dup_groups.values() for rid in v]
    return {
        "ids": flat_ids,
        "groups": [{"key_category": k[2], "ids": v} for k, v in dup_groups.items()],
        "note": f"{len(dup_groups)} duplicate groups on (system_rule, user_query, category), "
        f"{len(flat_ids)} rows total.",
        "blocking": False,
    }


def check_duplicate_ids(pairs):
    counts = Counter(r.get("id") for r in pairs)
    dups = {i: c for i, c in counts.items() if c > 1}
    return {
        "ids": sorted(dups.keys()),
        "note": (
            f"{len(dups)} id values appear more than once (should be 0)."
            if dups
            else "No duplicate ids."
        ),
        "blocking": True,
    }


def check_schema(pairs):
    ids_missing = []
    field_miss = Counter()
    for r in pairs:
        missing = [fld for fld in REQUIRED_FIELDS if not r.get(fld)]
        if missing:
            ids_missing.append(r.get("id"))
            for m in missing:
                field_miss[m] += 1
    return {
        "ids": ids_missing,
        "field_breakdown": dict(field_miss),
        "note": (
            f"{len(ids_missing)} rows missing >=1 required field. By field: {dict(field_miss)}"
            if ids_missing
            else "All rows have required fields."
        ),
        "blocking": True,
    }


CHECKS_NEEDING_ENGLISH = {"translation_drift"}


def run_all_checks(pairs, english_by_id, is_english):
    results = {}
    results["identical_contrastive_pair"] = check_identical_contrastive_pair(pairs)
    results["missing_pair_type"] = check_missing_pair_type(pairs)
    if not is_english:
        results["translation_drift"] = check_translation_drift(pairs, english_by_id)
    results["category_validity"] = check_category_validity(pairs)
    results["llm_judge_checker"] = check_llm_judge_checkers(pairs)
    results["missing_checker"] = check_missing_checker(pairs)
    results["custom_checker_dependency"] = check_custom_checker_dependency(pairs)
    results["duplicate_rows"] = check_duplicate_rows(pairs)
    results["duplicate_ids"] = check_duplicate_ids(pairs)
    results["schema_missing_fields"] = check_schema(pairs)
    return results


# ---------------------------------------------------------------------------
# Excel output -- one sheet per language, issues grouped, IDs listed per group.
# ---------------------------------------------------------------------------


def write_workbook(all_results, row_counts, en_row_count, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BLOCK_FILL = PatternFill("solid", fgColor="FCE4D6")  # blocking issue rows
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")  # clean rows
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")  # non-blocking issue rows
    BASE_FONT = Font(name="Arial", size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ---- SUMMARY sheet ----
    ws = wb.active
    ws.title = "SUMMARY"
    summary_headers = [
        "Language",
        "Rows",
        "Rows vs EN base",
        "Identical pair (1)",
        "Missing pair_type (2)",
        "Translation drift (3)",
        "Undoc. LLM-judge cats (4)",
        "Dup rows (6)",
        "Dup ids (7)",
        "Invalid category",
        "Missing checker",
        "Schema-missing",
        "Status",
    ]
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    r = 2
    for lang, results in all_results.items():
        n = row_counts[lang]
        drift = results.get("translation_drift", {}).get("ids", [])
        undoc = results["llm_judge_checker"]["categories_fully_llm_judge_UNDOCUMENTED"]
        vals = [
            lang,
            n,
            (n - en_row_count) if lang != "en" else 0,
            len(results["identical_contrastive_pair"]["ids"]),
            len(results["missing_pair_type"]["ids"]),
            len(drift),
            ", ".join(undoc) if undoc else "-",
            len(results["duplicate_rows"]["ids"]),
            len(results["duplicate_ids"]["ids"]),
            len(results["category_validity"]["ids"]),
            len(results["missing_checker"]["ids"]),
            len(results["schema_missing_fields"]["ids"]),
        ]
        blocking_hit = any(
            results[k].get("blocking") and results[k]["ids"] for k in results
        )
        vals.append("BLOCKING ISSUES" if blocking_hit else "clean")
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = BASE_FONT
            cell.alignment = WRAP
            if c == len(vals):
                cell.fill = BLOCK_FILL if blocking_hit else OK_FILL
        r += 1

    for c in range(1, len(summary_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 10
    ws.freeze_panes = "A2"

    # ---- per-language sheets ----
    ISSUE_ORDER = [
        ("identical_contrastive_pair", "1. Identical contrastive pair (zero signal)"),
        ("translation_drift", "3. Translation-introduced drift (vs EN)"),
        ("duplicate_ids", "7. Duplicate IDs"),
        ("category_validity", "Schema: invalid category value"),
        ("missing_checker", "Schema: missing checker"),
        ("schema_missing_fields", "Schema: missing required fields"),
        ("missing_pair_type", "2. Missing pair_type"),
        ("llm_judge_checker", "4. LLM-judge checker categories"),
        ("duplicate_rows", "6. Duplicate rows"),
        ("custom_checker_dependency", "5. Custom-checker dependency"),
    ]

    for lang, results in all_results.items():
        ws = wb.create_sheet(title=lang)
        headers = ["Issue", "Severity", "Count", "Note", "Affected row IDs (grouped)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP
        rr = 2
        for key, label in ISSUE_ORDER:
            if key not in results:
                continue
            res = results[key]
            ids = res.get("ids", [])
            blocking = res.get("blocking", False)
            severity = "BLOCKING" if blocking else "review"
            # ID text: for issues with sub-groups, show the split.
            id_text = format_ids_grouped(key, res)
            row_vals = [label, severity, len(ids), res.get("note", ""), id_text]
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(rr, c, v)
                cell.font = BASE_FONT
                cell.alignment = WRAP
            fill = OK_FILL if not ids else (BLOCK_FILL if blocking else WARN_FILL)
            for c in range(1, len(headers) + 1):
                ws.cell(rr, c).fill = fill
            rr += 1

        widths = [40, 12, 8, 55, 70]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    wb.save(out_path)


def format_ids_grouped(key, res):
    """Render affected IDs, showing sub-group splits where the check provides them,
    so the sheet reads as grouped triage rather than an undifferentiated ID dump."""

    def block(title, ids, cap=200):
        if not ids:
            return ""
        shown = ids[:cap]
        more = f"  (+{len(ids) - cap} more)" if len(ids) > cap else ""
        return f"[{title}: {len(ids)}]\n" + ", ".join(str(i) for i in shown) + more

    parts = []
    if key == "missing_pair_type":
        parts.append(
            block(
                "well-formed, metadata backfill only",
                res.get("ids_wellformed_metadata_only", []),
            )
        )
        parts.append(
            block(
                "also identical-pair (needs regen)",
                res.get("ids_also_identical_pair", []),
            )
        )
    elif key == "translation_drift":
        parts.append(
            block("contrastive pair collapsed", res.get("ids_collapsed_pair", []))
        )
        parts.append(block("pair_type dropped", res.get("ids_dropped_pairtype", [])))
    elif key == "llm_judge_checker":
        doc = res.get("categories_fully_llm_judge_documented", [])
        und = res.get("categories_fully_llm_judge_UNDOCUMENTED", [])
        parts.append(f"documented-intentional cats: {', '.join(doc) or 'none'}")
        parts.append(f"UNDOCUMENTED cats: {', '.join(und) or 'none'}")
        parts.append(block("rows in undocumented cats", res.get("ids", [])))
    elif key == "duplicate_rows":
        for g in res.get("groups", [])[:60]:
            parts.append(
                f"[{g['key_category']}] " + ", ".join(str(i) for i in g["ids"])
            )
    else:
        parts.append(block("ids", res.get("ids", [])))
    return "\n".join(p for p in parts if p)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument(
        "--data-dir", required=True, help="Directory containing full_dataset*.json"
    )
    ap.add_argument(
        "--out",
        default="validation_report.xlsx",
        help="Output .xlsx path (default: validation_report.xlsx)",
    )
    ap.add_argument(
        "--json-out", default=None, help="Optional machine-readable JSON report path"
    )
    ap.add_argument(
        "--ci",
        action="store_true",
        help="Exit non-zero if any BLOCKING issue is present (for CI gating)",
    )
    args = ap.parse_args()

    # Load English base first -- it's the reference for drift.
    en_path = os.path.join(args.data_dir, ENGLISH_FILENAME)
    if not os.path.exists(en_path):
        sys.exit(
            f"ERROR: English base not found at {en_path} -- it is required as the "
            f"reference for the translation-drift check."
        )
    en_pairs, _ = load_pairs(en_path)
    english_by_id = {r.get("id"): r for r in en_pairs}
    en_row_count = len(en_pairs)
    print(f"English base: {en_row_count} rows loaded from {ENGLISH_FILENAME}")

    all_results = {}
    row_counts = {}
    json_report = {}

    for lang, fname in LANGUAGE_FILES.items():
        path = os.path.join(args.data_dir, fname)
        if not os.path.exists(path):
            print(f"  WARNING: {lang}: file not found ({fname}) -- skipped")
            continue
        pairs, _ = load_pairs(path)
        row_counts[lang] = len(pairs)
        results = run_all_checks(pairs, english_by_id, is_english=(lang == "en"))
        all_results[lang] = results
        json_report[lang] = {
            "row_count": len(pairs),
            "issues": {
                k: {
                    kk: vv
                    for kk, vv in v.items()
                    if kk != "ids" or len(v["ids"]) <= 1000
                }
                for k, v in results.items()
            },
        }
        # concise console line
        block_hits = [
            k for k in results if results[k].get("blocking") and results[k]["ids"]
        ]
        status = f"BLOCKING: {', '.join(block_hits)}" if block_hits else "clean"
        drift_n = len(results.get("translation_drift", {}).get("ids", []))
        print(
            f"  {lang}: {len(pairs):>4} rows | identical-pair "
            f"{len(results['identical_contrastive_pair']['ids']):>4} | "
            f"drift {drift_n:>4} | missing-pair_type "
            f"{len(results['missing_pair_type']['ids']):>4} | {status}"
        )

    write_workbook(all_results, row_counts, en_row_count, args.out)
    print(f"\nWrote {args.out} ({len(all_results)} language sheets + SUMMARY)")

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(json_report, f, ensure_ascii=False, indent=2)
        print(f"Wrote machine-readable report to {args.json_out}")

    if args.ci:
        any_blocking = any(
            all_results[lang][k].get("blocking") and all_results[lang][k]["ids"]
            for lang in all_results
            for k in all_results[lang]
        )
        if any_blocking:
            print("\nCI GATE: FAILED -- blocking issues present.")
            sys.exit(1)
        print("\nCI GATE: PASSED.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
